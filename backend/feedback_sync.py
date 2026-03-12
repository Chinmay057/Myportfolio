"""
feedback_sync.py
================
Syncs Firebase Firestore 'feedbacks' collection with a local Excel file.

HOW IT WORKS
------------
1. Run this script → it exports all Firestore feedbacks to feedback_data.xlsx
2. Open feedback_data.xlsx, delete any row(s) you want removed
3. Save the Excel file and run the script again
4. The script detects missing document IDs and deletes them from Firestore
5. The website reflects the changes instantly (next page load)

SETUP
-----
1. Install dependencies:
       pip install -r requirements.txt

2. Download your Firebase service account key:
   Firebase Console → Project Settings → Service Accounts
   → Generate new private key → save as 'serviceAccountKey.json'
   in this same folder (backend/)

3. Run:
       python feedback_sync.py
"""

import os
import sys
import json
from datetime import datetime, timezone

try:
    import firebase_admin
    from firebase_admin import credentials, firestore
except ImportError:
    sys.exit("❌  Run: pip install firebase-admin")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("❌  Run: pip install openpyxl")

# ── Config ────────────────────────────────────────────────────────────────────
SERVICE_ACCOUNT_FILE = os.path.join(os.path.dirname(__file__), "serviceAccountKey.json")
EXCEL_FILE           = os.path.join(os.path.dirname(__file__), "feedback_data.xlsx")
COLLECTION_NAME      = "feedbacks"

COLUMNS = ["doc_id", "name", "email", "rating", "message", "submitted_at"]
HEADER_LABELS = ["Document ID (DO NOT EDIT)", "Name", "Email", "Rating (1-5)", "Message", "Submitted At"]


# ── Firebase Init ─────────────────────────────────────────────────────────────
def init_firebase():
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        sys.exit(
            f"❌  Service account key not found at:\n    {SERVICE_ACCOUNT_FILE}\n\n"
            "    Download it from Firebase Console → Project Settings → Service Accounts."
        )
    if not firebase_admin._apps:
        cred = credentials.Certificate(SERVICE_ACCOUNT_FILE)
        firebase_admin.initialize_app(cred)
    return firestore.client()


# ── Read Firestore ─────────────────────────────────────────────────────────────
def fetch_from_firestore(db):
    docs = db.collection(COLLECTION_NAME).order_by("createdAt", direction=firestore.Query.DESCENDING).stream()
    records = []
    for doc in docs:
        data = doc.to_dict()
        ts = data.get("createdAt")
        submitted = ""
        if ts:
            try:
                submitted = ts.strftime("%Y-%m-%d %H:%M:%S UTC")
            except Exception:
                pass
        records.append({
            "doc_id":       doc.id,
            "name":         data.get("name", ""),
            "email":        data.get("email", ""),
            "rating":       data.get("rating", ""),
            "message":      data.get("message", ""),
            "submitted_at": submitted,
        })
    return records


# ── Write Excel ────────────────────────────────────────────────────────────────
def write_excel(records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feedbacks"

    # Header style
    header_fill = PatternFill("solid", fgColor="3B1F5E")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="555555")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [30, 20, 30, 14, 60, 25]

    for col_idx, (label, width) in enumerate(zip(HEADER_LABELS, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, rec in enumerate(records, start=2):
        row_fill = PatternFill("solid", fgColor="1A1A2E" if row_idx % 2 == 0 else "12122A")
        for col_idx, key in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=rec[key])
            cell.font = Font(color="DDDDDD", size=10)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=(key == "message"))
            cell.border = border
        ws.row_dimensions[row_idx].height = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(EXCEL_FILE)
    print(f"✅  Saved {len(records)} records → {EXCEL_FILE}")


# ── Read Excel doc IDs ─────────────────────────────────────────────────────────
def read_excel_ids():
    if not os.path.exists(EXCEL_FILE):
        return None  # First run — no Excel yet
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # doc_id column
            ids.add(str(row[0]).strip())
    return ids


# ── Delete from Firestore ──────────────────────────────────────────────────────
def delete_missing(db, firestore_ids, excel_ids):
    to_delete = firestore_ids - excel_ids
    if not to_delete:
        print("✅  No deletions needed — all Firestore docs are present in Excel.")
        return
    for doc_id in to_delete:
        db.collection(COLLECTION_NAME).document(doc_id).delete()
        print(f"🗑️   Deleted Firestore doc: {doc_id}")
    print(f"\n✅  Deleted {len(to_delete)} record(s) from Firestore (and thus the website).")


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print("🔥  Connecting to Firebase Firestore…")
    db = init_firebase()

    print("📥  Fetching feedbacks from Firestore…")
    records = fetch_from_firestore(db)
    firestore_ids = {r["doc_id"] for r in records}

    # Check if Excel already exists → detect deletions
    excel_ids = read_excel_ids()
    if excel_ids is not None:
        print(f"📊  Excel has {len(excel_ids)} rows | Firestore has {len(firestore_ids)} docs")
        delete_missing(db, firestore_ids, excel_ids)
        # Re-fetch after deletions
        records = fetch_from_firestore(db)
    else:
        print("📋  First run — creating Excel file…")

    # Write updated Excel
    write_excel(records)
    print("\nDone! Open feedback_data.xlsx to manage feedbacks.")
    print("Delete rows → save → run this script again to sync deletions to the website.")


if __name__ == "__main__":
    main()
